"""
FastAPI backend for PDF Table Extraction using PaddleOCR
"""
import numpy as np
from pathlib import Path
from statistics import median
from typing import List, Optional
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font
import tempfile
import shutil
import traceback
import threading
import re
import time
import os
import json

# PDF and image processing
from pdf2image import convert_from_path
from PIL import ImageFilter

try:
    import cv2
except ImportError:
    cv2 = None

# PaddleOCR
OCR_IMPORT_ERROR: Optional[str] = None
OCR_INIT_ERROR: Optional[str] = None

try:
    from paddleocr import PaddleOCR
except ImportError as e:
    OCR_IMPORT_ERROR = f"{type(e).__name__}: {e}"
    print(f"Warning: Failed to import PaddleOCR: {OCR_IMPORT_ERROR}")
    PaddleOCR = None

# Initialize FastAPI app
app = FastAPI(title="PDF Table Extractor")

# Add CORS middleware to allow frontend requests
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Serve static frontend files
FRONTEND_DIR = Path(__file__).parent.parent / "frontend"
if FRONTEND_DIR.exists():
    app.mount("/static", StaticFiles(directory=FRONTEND_DIR), name="static")

# Create uploads directory if it doesn't exist
UPLOAD_DIR = Path(__file__).parent.parent / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

# Initialize PaddleOCR
ocr = None
OCR_INIT_STARTED = False
ocr_lock = threading.Lock()


def _clear_paddleocr_cache():
    cache_root = Path.home() / ".paddleocr"
    if cache_root.exists():
        shutil.rmtree(cache_root, ignore_errors=True)


def get_ocr():
    """
    Lazily initialize PaddleOCR so the web server can start immediately even if
    model downloads take time on first use.
    """
    global ocr, OCR_INIT_ERROR, OCR_INIT_STARTED

    if ocr is not None:
        return ocr

    if PaddleOCR is None:
        return None

    with ocr_lock:
        if ocr is not None:
            return ocr

        OCR_INIT_STARTED = True
        OCR_INIT_ERROR = None
        last_error = None

        for attempt in range(2):
            try:
                ocr = PaddleOCR(use_angle_cls=True, lang='en', show_log=False)
                OCR_INIT_ERROR = None
                break
            except Exception as e:
                last_error = e
                OCR_INIT_ERROR = f"{type(e).__name__}: {e}"
                print(f"Error initializing PaddleOCR (attempt {attempt + 1}/2): {OCR_INIT_ERROR}")
                traceback.print_exc()
                ocr = None

                error_text = str(e).lower()
                should_retry = attempt == 0 and any(
                    token in error_text
                    for token in ["incompleteread", "chunkedencodingerror", "connection broken"]
                )
                if should_retry:
                    _clear_paddleocr_cache()
                    time.sleep(1.5)
                    continue

                break

        if ocr is None and last_error is not None:
            message = str(last_error)
            if any(token in message.lower() for token in ["incompleteread", "chunkedencodingerror", "connection broken"]):
                OCR_INIT_ERROR = (
                    "PaddleOCR model download was interrupted. "
                    "Please retry once the network is stable."
                )

    return ocr


class TableDetector:
    """
    Detects and extracts tables from OCR results using clustering heuristic.
    Groups OCR results into rows/columns using Y-axis clustering and X-axis sorting.
    """
    
    def __init__(self, y_threshold=15, x_threshold=10, table_gap_multiplier=2.4):
        """
        Args:
            y_threshold: Vertical distance threshold for same row
            x_threshold: Horizontal distance threshold (not critical, used for spacing)
        """
        self.y_threshold = y_threshold
        self.x_threshold = x_threshold
        self.table_gap_multiplier = table_gap_multiplier
        self.debug_snapshots = []

    def _record_debug(self, stage, payload):
        self.debug_snapshots.append({
            "stage": stage,
            "payload": payload,
        })

    def _extract_items(self, ocr_result):
        items = []
        for item in ocr_result:
            bbox, (text, confidence) = item
            text = text.strip()
            if not text:
                continue

            y_coords = [point[1] for point in bbox]
            x_coords = [point[0] for point in bbox]

            items.append({
                "text": text,
                "center_y": (min(y_coords) + max(y_coords)) / 2,
                "center_x": (min(x_coords) + max(x_coords)) / 2,
                "min_y": min(y_coords),
                "max_y": max(y_coords),
                "x": min(x_coords),
                "max_x": max(x_coords),
                "bbox": bbox,
                "confidence": confidence,
                "height": max(y_coords) - min(y_coords),
                "width": max(x_coords) - min(x_coords),
            })
        return items

    def _cluster_numeric_positions(self, values, min_gap=8):
        if not values:
            return []

        values = sorted(values)
        clusters = [[values[0]]]
        for value in values[1:]:
            if value - clusters[-1][-1] <= min_gap:
                clusters[-1].append(value)
            else:
                clusters.append([value])
        return [int(round(sum(cluster) / len(cluster))) for cluster in clusters]

    def _find_table_regions(self, image_np):
        if cv2 is None:
            return []

        gray = cv2.cvtColor(image_np, cv2.COLOR_RGB2GRAY)
        binary = cv2.adaptiveThreshold(
            gray,
            255,
            cv2.ADAPTIVE_THRESH_MEAN_C,
            cv2.THRESH_BINARY_INV,
            15,
            8,
        )

        horizontal_kernel = cv2.getStructuringElement(
            cv2.MORPH_RECT,
            (max(20, image_np.shape[1] // 40), 1),
        )
        vertical_kernel = cv2.getStructuringElement(
            cv2.MORPH_RECT,
            (1, max(20, image_np.shape[0] // 50)),
        )

        horizontal_lines = cv2.morphologyEx(binary, cv2.MORPH_OPEN, horizontal_kernel)
        vertical_lines = cv2.morphologyEx(binary, cv2.MORPH_OPEN, vertical_kernel)
        table_mask = cv2.add(horizontal_lines, vertical_lines)

        contours, _ = cv2.findContours(table_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        regions = []
        page_area = image_np.shape[0] * image_np.shape[1]

        for contour in contours:
            x, y, w, h = cv2.boundingRect(contour)
            area = w * h
            if w < image_np.shape[1] * 0.45 or h < 180:
                continue
            if area < page_area * 0.05:
                continue
            regions.append({
                "x": x,
                "y": y,
                "w": w,
                "h": h,
                "mask": table_mask[y:y + h, x:x + w],
            })

        regions.sort(key=lambda region: (region["y"], region["x"]))
        return regions

    def _line_positions_from_mask(self, mask, axis):
        if axis == "horizontal":
            projection = mask.sum(axis=1)
            threshold = max(mask.shape[1] * 255 * 0.18, projection.max() * 0.18 if projection.size else 0)
        else:
            projection = mask.sum(axis=0)
            threshold = max(mask.shape[0] * 255 * 0.18, projection.max() * 0.18 if projection.size else 0)

        raw_positions = [idx for idx, value in enumerate(projection) if value >= threshold]
        return self._cluster_numeric_positions(raw_positions, min_gap=10)

    def _extract_grid(self, region):
        horizontal_positions = self._line_positions_from_mask(region["mask"], "horizontal")
        vertical_positions = self._line_positions_from_mask(region["mask"], "vertical")

        if len(horizontal_positions) < 3 or len(vertical_positions) < 3:
            return None

        absolute_rows = [region["y"] + pos for pos in horizontal_positions]
        absolute_cols = [region["x"] + pos for pos in vertical_positions]

        return {
            "rows": absolute_rows,
            "cols": absolute_cols,
            "bbox": (region["x"], region["y"], region["w"], region["h"]),
        }

    def _cell_text(self, items, left, top, right, bottom):
        margin_x = max(4, self.x_threshold)
        margin_y = max(4, self.y_threshold)
        cell_items = [
            item for item in items
            if left - margin_x <= item["center_x"] <= right + margin_x
            and top - margin_y <= item["center_y"] <= bottom + margin_y
        ]
        return self._join_fragments(sorted(cell_items, key=lambda item: (item["center_y"], item["x"])))

    def _dynamic_row_threshold(self, items):
        heights = [item["height"] for item in items if item["height"] > 0]
        if not heights:
            return self.y_threshold

        median_height = float(median(heights))
        return max(6.0, min(median_height * 0.42, median_height * 0.8))

    def _dynamic_column_threshold(self, items):
        widths = [item["width"] for item in items if item["width"] > 0]
        if not widths:
            return self.x_threshold * 2

        median_width = float(median(widths))
        return max(12.0, min(median_width * 0.55, 40.0))

    def _items_in_region(self, items, region, padding=8):
        left = region["x"] - padding
        right = region["x"] + region["w"] + padding
        top = region["y"] - padding
        bottom = region["y"] + region["h"] + padding

        return [
            item for item in items
            if left <= item["center_x"] <= right
            and top <= item["center_y"] <= bottom
        ]

    def _cluster_rows_dynamic(self, items):
        if not items:
            return []

        threshold = self._dynamic_row_threshold(items)
        sorted_items = sorted(items, key=lambda item: item["center_y"])
        row_groups = [[sorted_items[0]]]

        for item in sorted_items[1:]:
            current_row = row_groups[-1]
            current_center = median(entry["center_y"] for entry in current_row)
            if abs(item["center_y"] - current_center) <= threshold:
                current_row.append(item)
            else:
                row_groups.append([item])

        return [self._build_row(row_items) for row_items in row_groups]

    def _cluster_column_anchors(self, positions, threshold):
        if not positions:
            return []

        sorted_positions = sorted(set(positions))
        if len(sorted_positions) == 1:
            return [float(sorted_positions[0])]

        gaps = [
            current - previous
            for previous, current in zip(sorted_positions, sorted_positions[1:])
            if current - previous > 0
        ]
        if gaps:
            gap_threshold = max(8.0, float(median(gaps)) * 1.5)
        else:
            gap_threshold = max(8.0, float(threshold or self.x_threshold))

        clusters = [[sorted_positions[0]]]
        for position in sorted_positions[1:]:
            if position - clusters[-1][-1] <= gap_threshold:
                clusters[-1].append(position)
            else:
                clusters.append([position])

        return [float(median(cluster)) for cluster in clusters]

    def _normalized_row_signature(self, row):
        return tuple(" ".join(str(cell).strip().split()) for cell in row)

    def _merge_sparse_row_values(self, values):
        merged_parts = []
        for value in values:
            cleaned = " ".join(value.strip().split())
            if cleaned and (not merged_parts or merged_parts[-1] != cleaned):
                merged_parts.append(cleaned)
        return " ".join(merged_parts).strip()

    def _merge_sparse_columns(self, rows, anchors=None, sparse_threshold=0.3, dense_threshold=0.55):
        if not rows:
            return rows, anchors or []

        column_count = max((len(row) for row in rows), default=0)
        if column_count < 2:
            return rows, anchors or []

        row_count = len(rows)
        fill_ratios = []
        numeric_ratios = []
        for col_idx in range(column_count):
            non_empty_values = [
                row[col_idx].strip()
                for row in rows
                if col_idx < len(row) and isinstance(row[col_idx], str) and row[col_idx].strip()
            ]
            filled = len(non_empty_values)
            fill_ratios.append(filled / row_count if row_count else 0.0)
            numeric_ratios.append(
                (sum(1 for value in non_empty_values if self._looks_numeric(value)) / filled)
                if filled else 0.0
            )

        merged_rows = []
        merged_anchors = []
        col_idx = 0

        while col_idx < column_count:
            if fill_ratios[col_idx] >= sparse_threshold:
                merged_rows.append([
                    row[col_idx] if col_idx < len(row) else ""
                    for row in rows
                ])
                if anchors and col_idx < len(anchors):
                    merged_anchors.append(anchors[col_idx])
                col_idx += 1
                continue

            run_start = col_idx
            while col_idx < column_count and fill_ratios[col_idx] < sparse_threshold:
                col_idx += 1
            run_end = col_idx

            if run_end - run_start < 2:
                merged_rows.append([
                    row[run_start] if run_start < len(row) else ""
                    for row in rows
                ])
                if anchors and run_start < len(anchors):
                    merged_anchors.append(anchors[run_start])
                continue

            if any(numeric_ratios[idx] > 0.35 for idx in range(run_start, run_end)):
                for idx in range(run_start, run_end):
                    merged_rows.append([
                        row[idx] if idx < len(row) else ""
                        for row in rows
                    ])
                    if anchors and idx < len(anchors):
                        merged_anchors.append(anchors[idx])
                continue

            combined_coverage = sum(
                1
                for row in rows
                if any(
                    idx < len(row) and isinstance(row[idx], str) and row[idx].strip()
                    for idx in range(run_start, run_end)
                )
            ) / row_count if row_count else 0.0

            if combined_coverage < dense_threshold:
                for idx in range(run_start, run_end):
                    merged_rows.append([
                        row[idx] if idx < len(row) else ""
                        for row in rows
                    ])
                    if anchors and idx < len(anchors):
                        merged_anchors.append(anchors[idx])
                continue

            merged_column = []
            for row in rows:
                merged_column.append(
                    self._merge_sparse_row_values([
                        row[idx]
                        for idx in range(run_start, min(run_end, len(row)))
                        if isinstance(row[idx], str) and row[idx].strip()
                    ])
                )
            merged_rows.append(merged_column)
            if anchors:
                run_anchors = [anchors[idx] for idx in range(run_start, min(run_end, len(anchors)))]
                merged_anchors.append(float(median(run_anchors)) if run_anchors else 0.0)

        normalized_rows = [
            [column[row_idx] for column in merged_rows]
            for row_idx in range(row_count)
        ] if merged_rows else rows

        return normalized_rows, merged_anchors

    def _remove_duplicate_rows(self, rows):
        deduped_rows = []
        seen = set()

        for row in rows:
            signature = self._normalized_row_signature(row)
            if not any(signature):
                continue
            if signature in seen:
                continue
            seen.add(signature)
            deduped_rows.append(row)

        return deduped_rows

    def _row_fill_count(self, row):
        return sum(1 for cell in row if isinstance(cell, str) and cell.strip())

    def _row_fill_ratio(self, row):
        column_count = max(1, len(row))
        return self._row_fill_count(row) / column_count

    def _row_is_non_numeric(self, row):
        filled = [cell.strip() for cell in row if isinstance(cell, str) and cell.strip()]
        return bool(filled) and all(not self._looks_numeric(cell) for cell in filled)

    def _merge_header_rows(self, rows):
        if not rows:
            return rows, 0

        fill_counts = [self._row_fill_count(row) for row in rows]
        populated_counts = [count for count in fill_counts if count > 0]
        if not populated_counts:
            return rows, 0

        median_fill = float(median(populated_counts))
        header_zone = []
        max_header_rows = min(4, max(1, len(rows) - 1))

        for idx, row in enumerate(rows[:max_header_rows]):
            fill_count = fill_counts[idx]
            if fill_count == 0:
                break

            sparse_like_header = fill_count <= max(1.0, median_fill * 0.7)
            textual_row = self._row_is_non_numeric(row)
            if idx == 0 and textual_row:
                header_zone.append(row)
                continue
            if sparse_like_header and textual_row:
                header_zone.append(row)
                continue
            break

        if not header_zone:
            inferred = self._infer_header_row_count(rows)
            return rows, inferred

        column_count = max(len(row) for row in rows)
        merged_header = []
        for col_idx in range(column_count):
            parts = []
            for row in header_zone:
                if col_idx >= len(row):
                    continue
                value = row[col_idx].strip()
                if value and (not parts or parts[-1] != value):
                    parts.append(value)
            merged_header.append(" ".join(parts).strip())

        remaining_rows = rows[len(header_zone):]
        return [merged_header] + remaining_rows, 1

    def _classify_body_rows(self, rows):
        row_types = []
        for row in rows:
            if self._row_fill_ratio(row) < 0.4 and self._row_is_non_numeric(row):
                row_types.append("section")
            else:
                row_types.append("data")
        return row_types

    def _is_section_like_row(self, row):
        return self._row_fill_ratio(row) < 0.4 and self._row_is_non_numeric(row)

    def _is_caption_like_row(self, row):
        filled = [cell.strip() for cell in row if isinstance(cell, str) and cell.strip()]
        if not filled or len(filled) > 2:
            return False

        joined = " ".join(filled).lower()
        return (
            joined.startswith("table ")
            or joined.startswith("figure ")
            or joined.startswith("appendix ")
            or joined.startswith("exhibit ")
        )

    def _strip_leading_caption_rows(self, rows):
        stripped = list(rows)
        while stripped and self._is_caption_like_row(stripped[0]):
            stripped.pop(0)
        return stripped

    def _dedupe_rows_with_types(self, rows, row_types):
        deduped_rows = []
        deduped_types = []
        seen = set()

        for row, row_type in zip(rows, row_types):
            signature = self._normalized_row_signature(row)
            if not any(signature):
                continue
            if signature in seen:
                continue
            seen.add(signature)
            deduped_rows.append(row)
            deduped_types.append(row_type)

        return deduped_rows, deduped_types

    def _build_row_grid(self, rows):
        if not rows:
            return [], []

        all_positions = [
            item["x"]
            for row in rows
            for item in row["items"]
        ]
        if not all_positions:
            return [], []

        row_items = [item for row in rows for item in row["items"]]
        anchors = self._cluster_column_anchors(all_positions, self._dynamic_column_threshold(row_items))
        if not anchors:
            return [], []

        table_rows = []
        for row in rows:
            values = [""] * len(anchors)
            occupied = set()

            for item in sorted(row["items"], key=lambda entry: entry["x"]):
                ranked = sorted(
                    range(len(anchors)),
                    key=lambda idx: (
                        abs(anchors[idx] - item["x"]),
                        abs(anchors[idx] - item["center_x"]),
                    ),
                )
                col_idx = ranked[0]
                if col_idx in occupied:
                    next_free = next((idx for idx in ranked if idx not in occupied), col_idx)
                    col_idx = next_free

                values[col_idx] = (
                    f"{values[col_idx]} {item['text']}".strip()
                    if values[col_idx]
                    else item["text"]
                )
                occupied.add(col_idx)

            table_rows.append(values)

        visible_columns = [
            idx for idx in range(len(anchors))
            if any((row[idx] if idx < len(row) else "").strip() for row in table_rows)
        ]
        normalized_rows = [
            [row[idx] for idx in visible_columns]
            for row in table_rows
        ]
        normalized_anchors = [anchors[idx] for idx in visible_columns]

        normalized_rows, normalized_anchors = self._merge_sparse_columns(normalized_rows, normalized_anchors)

        return normalized_rows, normalized_anchors

    def _split_grid_into_tables(self, grid_rows, min_cols=2, min_rows=2):
        if not grid_rows:
            return []

        fill_counts = [sum(1 for cell in row if cell.strip()) for row in grid_rows]
        non_zero_counts = [count for count in fill_counts if count > 0]
        if not non_zero_counts:
            return []

        median_fill = float(median(non_zero_counts))
        threshold = max(min_cols, int(round(median_fill * 0.4)))

        tables = []
        current = []
        consecutive_sparse = 0
        for row, fill_count in zip(grid_rows, fill_counts):
            if fill_count >= threshold:
                current.append(row)
                consecutive_sparse = 0
            else:
                current.append(row)
                consecutive_sparse += 1
                if consecutive_sparse >= 2:
                    table_candidate = current[:-2]
                    if len(table_candidate) >= min_rows:
                        tables.append(table_candidate)
                    current = current[-2:]
                    consecutive_sparse = 2

        trailing_sparse = consecutive_sparse
        if trailing_sparse >= 2:
            current = current[:-2]

        if len(current) >= min_rows:
            tables.append(current)

        return tables

    def _infer_header_merges_from_rows(self, rows, header_row_count):
        merges = []
        for row_idx in range(min(header_row_count, len(rows))):
            row = rows[row_idx]
            col_idx = 0
            while col_idx < len(row):
                value = row[col_idx].strip()
                if not value:
                    col_idx += 1
                    continue

                end_idx = col_idx
                while end_idx + 1 < len(row) and not row[end_idx + 1].strip():
                    end_idx += 1

                if end_idx > col_idx:
                    merges.append({
                        "row": row_idx,
                        "start_col": col_idx,
                        "end_col": end_idx,
                    })
                col_idx = end_idx + 1

        return merges

    def _assemble_table(self, rows, body_builder, fallback_merges=None):
        if not rows:
            return None

        rows = self._strip_leading_caption_rows(rows)
        if not rows:
            return None
        rows = self._remove_duplicate_rows(rows)
        rows, header_row_count = self._merge_header_rows(rows)
        header_rows = rows[:header_row_count]
        body_rows = body_builder(rows[header_row_count:])
        header_rows, body_rows = self._realign_table_to_layout(header_rows, body_rows)

        row_types = (["header"] * len(header_rows)) + self._classify_body_rows(body_rows)
        final_rows, row_types = self._dedupe_rows_with_types(header_rows + body_rows, row_types)
        header_row_count = 0
        for row_type in row_types:
            if row_type == "header":
                header_row_count += 1
            else:
                break

        return {
            "rows": final_rows,
            "row_types": row_types,
            "header_row_count": header_row_count,
            "merges": self._infer_header_merges_from_rows(final_rows, header_row_count) or (fallback_merges or []),
            "row_count": len(final_rows),
            "column_count": max((len(row) for row in final_rows), default=0),
        }

    def _finalize_table(self, rows):
        if not rows:
            return None

        normalized_rows = self._normalize_organization_column(rows)
        return self._assemble_table(normalized_rows, self._consolidate_body_rows)

    def _build_table_from_row_grid(self, items):
        dynamic_rows = self._cluster_rows_dynamic(items)
        trimmed_rows = self._trim_non_table_rows(self._remove_artifact_rows(dynamic_rows))
        if len(trimmed_rows) < 2:
            return []

        grid_rows, _ = self._build_row_grid(trimmed_rows)
        table_blocks = self._split_grid_into_tables(grid_rows)
        tables = []
        for block in table_blocks:
            table = self._finalize_table(block)
            if table:
                tables.append(table)
        return tables

    def _build_table_from_grid(self, grid, items):
        rows = grid["rows"]
        cols = grid["cols"]
        if len(rows) < 3 or len(cols) < 3:
            return None

        table_rows = []
        for row_idx in range(len(rows) - 1):
            row_values = []
            top = rows[row_idx]
            bottom = rows[row_idx + 1]
            if bottom - top < 10:
                continue

            for col_idx in range(len(cols) - 1):
                left = cols[col_idx]
                right = cols[col_idx + 1]
                if right - left < 10:
                    continue
                row_values.append(self._cell_text(items, left, top, right, bottom))

            if any(value.strip() for value in row_values):
                table_rows.append(row_values)

        if len(table_rows) < 3:
            return None

        table_rows, _ = self._merge_sparse_columns(table_rows)
        return self._assemble_table(table_rows, self._consolidate_grid_body_rows)

    def _table_rows_for_scoring(self, table):
        if isinstance(table, dict):
            return table.get("rows", [])
        return table or []

    def _score_table_set(self, tables):
        if not tables:
            return float("-inf")

        score = 0.0
        for table in tables:
            rows = self._table_rows_for_scoring(table)
            row_count = len(rows)
            column_count = max((len(row) for row in rows), default=0)
            if row_count == 0 or column_count == 0:
                continue

            non_empty_cells = sum(
                1
                for row in rows
                for cell in row
                if isinstance(cell, str) and cell.strip()
            )
            cell_capacity = row_count * column_count
            fill_ratio = non_empty_cells / cell_capacity if cell_capacity else 0.0
            header_row_count = table.get("header_row_count", 0) if isinstance(table, dict) else 0

            score += non_empty_cells * 4.0
            score += row_count * 2.0
            score += min(column_count, 12)
            score += min(fill_ratio, 1.0) * 10.0
            if header_row_count > 0:
                score += 6.0
            if fill_ratio < 0.45:
                score -= 10.0

        score -= max(0, len(tables) - 1) * 8.0
        return score

    def _select_best_tables(self, *candidate_sets):
        available_sets = [tables for tables in candidate_sets if tables]
        if not available_sets:
            return []

        return max(available_sets, key=self._score_table_set)

    def _consolidate_grid_body_rows(self, rows):
        if not rows:
            return rows

        rows = self._merge_continuation_rows(rows)
        consolidated = []
        current = None

        for row in rows:
            normalized = list(row)
            row_number = normalized[0].strip() if normalized else ""
            numeric_count = sum(1 for value in normalized[2:] if value.strip() and self._looks_numeric(value))
            is_section_row = self._is_section_like_row(normalized)

            if is_section_row:
                if current:
                    consolidated.append(current)
                    current = None
                consolidated.append(normalized)
                continue

            if row_number:
                if current:
                    consolidated.append(current)
                current = normalized
                continue

            if current and numeric_count == 0:
                for idx, value in enumerate(normalized):
                    if value.strip():
                        current[idx] = f"{current[idx]} {value}".strip() if current[idx].strip() else value
                continue

            if current:
                consolidated.append(current)
            current = normalized

        if current:
            consolidated.append(current)

        return consolidated

    def _cluster_rows(self, items):
        items_sorted_by_y = sorted(items, key=lambda item: item["center_y"])
        rows = []
        current_row = [items_sorted_by_y[0]]

        for item in items_sorted_by_y[1:]:
            if abs(item["center_y"] - current_row[0]["center_y"]) <= self.y_threshold:
                current_row.append(item)
            else:
                rows.append(self._build_row(current_row))
                current_row = [item]

        rows.append(self._build_row(current_row))
        return rows

    def _build_row(self, row_items):
        row_items.sort(key=lambda item: item["x"])
        return {
            "cells": [item["text"] for item in row_items],
            "items": row_items,
            "item_count": len(row_items),
            "min_y": min(item["min_y"] for item in row_items),
            "max_y": max(item["max_y"] for item in row_items),
            "height": max(item["max_y"] for item in row_items) - min(item["min_y"] for item in row_items),
            "min_x": min(item["x"] for item in row_items),
            "max_x": max(item["max_x"] for item in row_items),
        }

    def _table_gap_threshold(self, rows):
        heights = [row["height"] for row in rows if row["height"] > 0]
        base_height = median(heights) if heights else self.y_threshold
        return max(self.y_threshold * 2, base_height * self.table_gap_multiplier)

    def _rows_belong_to_same_table(self, previous_row, current_row, gap_threshold):
        gap = current_row["min_y"] - previous_row["max_y"]
        horizontal_overlap = min(previous_row["max_x"], current_row["max_x"]) - max(previous_row["min_x"], current_row["min_x"])

        if gap <= gap_threshold:
            return True

        if horizontal_overlap > 0 and gap <= gap_threshold * 1.35:
            return True

        return False

    def _is_probable_table(self, table_rows):
        if len(table_rows) < 2:
            return False

        max_columns = max(row["item_count"] for row in table_rows)
        multi_cell_rows = sum(1 for row in table_rows if row["item_count"] >= 2)

        return max_columns >= 2 and multi_cell_rows >= 2

    def _split_tables(self, rows):
        if not rows:
            return []

        gap_threshold = self._table_gap_threshold(rows)
        table_groups = []
        current_group = [rows[0]]

        for row in rows[1:]:
            previous_row = current_group[-1]
            if self._rows_belong_to_same_table(previous_row, row, gap_threshold):
                current_group.append(row)
            else:
                if self._is_probable_table(current_group):
                    table_groups.append(current_group)
                current_group = [row]

        if self._is_probable_table(current_group):
            table_groups.append(current_group)

        return table_groups

    def _trim_non_table_rows(self, rows):
        trimmed_rows = list(rows)

        while len(trimmed_rows) > 1 and trimmed_rows[0]["item_count"] <= 1:
            later_multi_cell_rows = any(row["item_count"] >= 2 for row in trimmed_rows[1:])
            if not later_multi_cell_rows:
                break
            trimmed_rows.pop(0)

        while len(trimmed_rows) > 1 and trimmed_rows[-1]["item_count"] <= 1:
            earlier_multi_cell_rows = any(row["item_count"] >= 2 for row in trimmed_rows[:-1])
            if not earlier_multi_cell_rows:
                break
            trimmed_rows.pop()

        return trimmed_rows

    def _is_artifact_text(self, text):
        normalized = " ".join(text.lower().split())
        if not normalized:
            return True

        artifact_patterns = [
            r"^table \d+$",
            r"^download as excel$",
            r"^pdf table extractor$",
            r"^localhost:\d+$",
            r"^\d{1,2}/\d{1,2}/\d{2,4},",
            r"^\d+/\d+$",
        ]
        return any(re.search(pattern, normalized) for pattern in artifact_patterns)

    def _remove_artifact_rows(self, rows):
        cleaned_rows = []
        for row in rows:
            texts = [item["text"] for item in row["items"] if item["text"].strip()]
            if texts and all(self._is_artifact_text(text) for text in texts):
                continue
            cleaned_rows.append(row)
        return cleaned_rows

    def _remove_artifact_items(self, items):
        return [item for item in items if not self._is_artifact_text(item["text"])]

    def _cluster_positions(self, values, threshold):
        if not values:
            return []

        sorted_values = sorted(values)
        clusters = [[sorted_values[0]]]
        for value in sorted_values[1:]:
            if abs(value - median(clusters[-1])) <= threshold:
                clusters[-1].append(value)
            else:
                clusters.append([value])
        return [median(cluster) for cluster in clusters]

    def _cluster_items_by_axis(self, items, key, threshold):
        if not items:
            return []

        sorted_items = sorted(items, key=lambda item: item[key])
        clusters = [[sorted_items[0]]]

        for item in sorted_items[1:]:
            cluster_center = median([cluster_item[key] for cluster_item in clusters[-1]])
            if abs(item[key] - cluster_center) <= threshold:
                clusters[-1].append(item)
            else:
                clusters.append([item])

        return clusters

    def _is_measure_value(self, text):
        stripped = text.strip()
        if not self._looks_numeric(stripped):
            return False

        compact = stripped.replace(",", "").replace("%", "").replace("(", "").replace(")", "")
        return len(compact) > 3 or "." in compact

    def _is_row_index_value(self, text):
        stripped = text.strip()
        compact = stripped.replace(",", "")
        return compact.isdigit() and len(compact) <= 3

    def _infer_metric_columns(self, items):
        numeric_items = [item for item in items if self._is_measure_value(item["text"])]
        if len(numeric_items) < 8:
            return []

        widths = [item["width"] for item in numeric_items if item["width"] > 0]
        threshold = max(self.x_threshold * 3, (median(widths) if widths else 0) * 0.8)
        clusters = self._cluster_items_by_axis(numeric_items, "center_x", threshold)
        metric_columns = []

        for cluster in clusters:
            center_x = median([item["center_x"] for item in cluster])
            metric_columns.append({
                "center_x": center_x,
                "items": cluster,
                "count": len(cluster),
            })

        metric_columns = [column for column in metric_columns if column["count"] >= 3]
        metric_columns.sort(key=lambda column: column["center_x"])

        if len(metric_columns) <= 4:
            return metric_columns

        return sorted(metric_columns, key=lambda column: (-column["count"], column["center_x"]))[:4]

    def _assign_metric_index(self, item, metric_columns):
        if not metric_columns:
            return None

        ranked = sorted(
            range(len(metric_columns)),
            key=lambda idx: abs(metric_columns[idx]["center_x"] - item["center_x"])
        )
        return ranked[0]

    def _build_numeric_row_anchors(self, items, metric_columns):
        numeric_items = [item for item in items if self._is_measure_value(item["text"])]
        if not numeric_items or not metric_columns:
            return []

        heights = [item["height"] for item in numeric_items if item["height"] > 0]
        threshold = max(self.y_threshold * 1.8, (median(heights) if heights else self.y_threshold) * 1.4)
        y_clusters = self._cluster_items_by_axis(numeric_items, "center_y", threshold)
        anchors = []

        for cluster in y_clusters:
            metric_indexes = {}
            for item in cluster:
                metric_idx = self._assign_metric_index(item, metric_columns)
                if metric_idx is None:
                    continue
                current = metric_indexes.get(metric_idx)
                if current is None or abs(item["center_x"] - metric_columns[metric_idx]["center_x"]) < abs(current["center_x"] - metric_columns[metric_idx]["center_x"]):
                    metric_indexes[metric_idx] = item

            if len(metric_indexes) >= min(3, len(metric_columns)):
                anchors.append({
                    "center_y": median([item["center_y"] for item in cluster]),
                    "items": cluster,
                    "metric_items": metric_indexes,
                    "min_y": min(item["min_y"] for item in cluster),
                    "max_y": max(item["max_y"] for item in cluster),
                })

        anchors.sort(key=lambda anchor: anchor["center_y"])
        return anchors

    def _split_anchor_groups(self, anchors):
        if not anchors:
            return []

        if len(anchors) == 1:
            return [anchors]

        gaps = [
            anchors[idx]["center_y"] - anchors[idx - 1]["center_y"]
            for idx in range(1, len(anchors))
        ]
        gap_threshold = median(gaps) * 1.9 if gaps else self.y_threshold * 4
        gap_threshold = max(gap_threshold, self.y_threshold * 4)

        groups = [[anchors[0]]]
        for anchor in anchors[1:]:
            previous = groups[-1][-1]
            gap = anchor["center_y"] - previous["center_y"]
            if gap > gap_threshold:
                groups.append([anchor])
            else:
                groups[-1].append(anchor)
        return groups

    def _text_band_items(self, items, top, bottom):
        return [
            item for item in items
            if item["center_y"] >= top and item["center_y"] <= bottom
        ]

    def _join_fragments(self, items):
        if not items:
            return ""

        row_groups = self._cluster_items_by_axis(items, "center_y", max(self.y_threshold, 12))
        lines = []
        for row in row_groups:
            row = sorted(row, key=lambda item: item["x"])
            line = " ".join(item["text"].strip() for item in row if item["text"].strip())
            if line:
                lines.append(line)
        return " ".join(lines).strip()

    def detect_tables_from_image(self, image_np, ocr_instance):
        if ocr_instance is None:
            return []

        ocr_result = normalize_ocr_result(ocr_instance.ocr(image_np, cls=True))
        items = self._remove_artifact_items(self._extract_items(ocr_result))
        self._record_debug("ocr_items", {
            "count": len(items),
            "sample": items[:20],
        })
        if not items:
            return []

        region_tables = []
        if cv2 is not None:
            regions = self._find_table_regions(image_np)
            for region in regions:
                region_items = self._items_in_region(items, region)
                row_grid_tables = self._build_table_from_row_grid(region_items)
                if row_grid_tables:
                    region_tables.extend(row_grid_tables)
                    continue

                grid = self._extract_grid(region)
                if not grid:
                    continue

                table = self._build_table_from_grid(grid, region_items)
                if table:
                    region_tables.append(table)

        page_tables = self._build_table_from_row_grid(items)
        self._record_debug("page_row_grid", page_tables)
        structured_tables = self._extract_structured_tables(items)
        self._record_debug("structured_tables", structured_tables)
        fallback_tables = self._detect_tables_fallback(items)
        self._record_debug("fallback_tables", fallback_tables)
        selected_tables = self._select_best_tables(region_tables, page_tables, structured_tables, fallback_tables)
        self._record_debug("selected_tables", selected_tables)
        return self._normalize_selected_tables(selected_tables)

    def _infer_table_x_bounds(self, anchors, metric_columns):
        right_edge = max(
            item["max_x"]
            for anchor in anchors
            for item in anchor["metric_items"].values()
        )
        left_edge = min(metric_columns[0]["center_x"] - max(120, self.x_threshold * 8), right_edge - 200)
        return left_edge, right_edge

    def _infer_header_rows_from_bands(self, header_items, top, first_anchor_top, left_edge, first_metric_left, column_count):
        header_candidates = [
            item for item in header_items
            if item["center_y"] >= top
            and item["center_y"] <= first_anchor_top
            and item["center_x"] >= left_edge - 40
        ]
        if not header_candidates:
            return []

        row_groups = self._cluster_items_by_axis(header_candidates, "center_y", max(self.y_threshold, 14))
        header_rows = []
        for row in row_groups:
            values = [""] * column_count
            for item in sorted(row, key=lambda entry: entry["x"]):
                target_idx = 1 if item["center_x"] < first_metric_left else 2
                values[target_idx] = f"{values[target_idx]} {item['text']}".strip() if values[target_idx] else item["text"]
            if any(value.strip() for value in values):
                header_rows.append(values)

        if header_rows:
            for row in header_rows:
                if any("no" == cell.strip().lower().rstrip(".") for cell in row if cell):
                    row[0] = next((cell for cell in row if cell.strip().lower().startswith("no")), row[0])
                if not row[1]:
                    possible_text = next((cell for cell in row if cell and not self._looks_numeric(cell)), "")
                    row[1] = possible_text
        return header_rows

    def _build_rows_from_anchor_group(self, anchors, items, metric_columns, previous_bottom, next_top):
        if not anchors:
            return None

        first_metric_left = min(column["center_x"] for column in metric_columns)
        left_edge, right_edge = self._infer_table_x_bounds(anchors, metric_columns)
        bands = []
        for idx, anchor in enumerate(anchors):
            top = previous_bottom if idx == 0 else (anchors[idx - 1]["center_y"] + anchor["center_y"]) / 2
            bottom = next_top if idx == len(anchors) - 1 else (anchor["center_y"] + anchors[idx + 1]["center_y"]) / 2
            bands.append((top, bottom))

        header_top = max(previous_bottom, anchors[0]["center_y"] - (anchors[1]["center_y"] - anchors[0]["center_y"]) * 1.6) if len(anchors) > 1 else previous_bottom
        header_rows = self._infer_header_rows_from_bands(
            items,
            header_top,
            bands[0][0],
            left_edge,
            first_metric_left,
            2 + len(metric_columns),
        )

        body_rows = []
        for anchor, (top, bottom) in zip(anchors, bands):
            band_items = self._text_band_items(items, top, bottom)
            row = [""] * (2 + len(metric_columns))

            left_items = [
                item for item in band_items
                if item["center_x"] < first_metric_left - 20
            ]
            number_items = [item for item in left_items if self._is_row_index_value(item["text"])]
            if number_items:
                number_items.sort(key=lambda item: item["x"])
                row[0] = number_items[0]["text"].strip()

            text_items = [
                item for item in left_items
                if item not in number_items and not self._is_measure_value(item["text"])
            ]
            row[1] = self._join_fragments(text_items)

            for metric_idx, metric_item in anchor["metric_items"].items():
                row[2 + metric_idx] = metric_item["text"].strip()

            if row[1] or any(row[2:]):
                body_rows.append(row)

        if not body_rows:
            return None

        header_rows, body_rows = self._realign_table_to_layout(header_rows, body_rows)
        final_rows = header_rows + body_rows
        return {
            "rows": final_rows,
            "header_row_count": len(header_rows),
            "merges": self._infer_header_merges_from_rows(final_rows, len(header_rows)),
        }

    def _extract_structured_tables(self, items):
        cleaned_items = self._remove_artifact_items(items)
        metric_columns = self._infer_metric_columns(cleaned_items)
        if len(metric_columns) < 4:
            return []

        metric_columns = sorted(metric_columns, key=lambda column: column["center_x"])
        anchors = self._build_numeric_row_anchors(cleaned_items, metric_columns)
        if len(anchors) < 2:
            return []

        anchor_groups = self._split_anchor_groups(anchors)
        tables = []
        for group_idx, group in enumerate(anchor_groups):
            previous_bottom = 0 if group_idx == 0 else (anchor_groups[group_idx - 1][-1]["center_y"] + group[0]["center_y"]) / 2
            next_top = float("inf") if group_idx == len(anchor_groups) - 1 else (group[-1]["center_y"] + anchor_groups[group_idx + 1][0]["center_y"]) / 2
            table = self._build_rows_from_anchor_group(group, cleaned_items, metric_columns, previous_bottom, next_top)
            if table:
                tables.append(table)
        return tables

    def _candidate_data_rows(self, rows):
        if not rows:
            return []

        max_columns = max(row["item_count"] for row in rows)
        return [row for row in rows if row["item_count"] >= max(2, max_columns - 1)]

    def _infer_column_centers(self, rows):
        candidate_rows = self._candidate_data_rows(rows)
        source_rows = candidate_rows or rows
        items = [item for row in source_rows for item in row["items"]]
        if not items:
            return []

        widths = [item["width"] for item in items if item["width"] > 0]
        median_width = median(widths) if widths else 0
        threshold = max(self.x_threshold * 2, median_width * 0.45 if median_width else 0)
        centers = [item["center_x"] for item in items]
        return self._cluster_positions(centers, threshold)

    def _build_column_bounds(self, column_centers):
        if not column_centers:
            return []

        bounds = []
        for idx, center in enumerate(column_centers):
            left = float("-inf") if idx == 0 else (column_centers[idx - 1] + center) / 2
            right = float("inf") if idx == len(column_centers) - 1 else (center + column_centers[idx + 1]) / 2
            bounds.append((left, right))
        return bounds

    def _find_column_index(self, item, column_centers, column_bounds, used_indexes):
        ranked_indexes = []
        for idx, center in enumerate(column_centers):
            left, right = column_bounds[idx]
            overlap = min(item["max_x"], right) - max(item["x"], left)
            distance = abs(center - item["center_x"])
            ranked_indexes.append((overlap, -distance, idx))

        ranked_indexes.sort(reverse=True)
        for overlap, _, idx in ranked_indexes:
            if idx not in used_indexes and overlap > 0:
                return idx

        fallback_indexes = sorted(
            range(len(column_centers)),
            key=lambda idx: abs(column_centers[idx] - item["center_x"])
        )
        for idx in fallback_indexes:
            if idx not in used_indexes:
                return idx
        return fallback_indexes[0] if fallback_indexes else 0

    def _split_leading_number_cell(self, row, expected_columns):
        if len(row) < 2 or len(row) != expected_columns:
            return row

        first_value = row[0].strip()
        if not first_value:
            return row

        match = re.match(r"^(\d+)\s+(.+)$", first_value)
        if not match:
            return row

        if row[1].strip():
            return row

        row_copy = list(row)
        row_copy[0] = match.group(1)
        row_copy[1] = match.group(2)
        return row_copy

    def _normalize_organization_column(self, rows):
        if not rows:
            return rows

        expected_columns = max(len(row) for row in rows)
        return [self._split_leading_number_cell(row, expected_columns) for row in rows]

    def _integer_like(self, value):
        compact = value.replace(",", "").strip()
        return bool(compact) and compact.isdigit()

    def _column_numeric_ratio(self, rows, column_idx):
        non_empty = [row[column_idx].strip() for row in rows if column_idx < len(row) and row[column_idx].strip()]
        if not non_empty:
            return 0
        numeric = sum(1 for value in non_empty if self._looks_numeric(value))
        return numeric / len(non_empty)

    def _infer_body_layout(self, body_rows):
        if not body_rows:
            return None

        column_count = max(len(row) for row in body_rows)
        numeric_columns = [
            idx for idx in range(column_count)
            if self._column_numeric_ratio(body_rows, idx) >= 0.55
        ]
        if not numeric_columns:
            return None

        first_numeric_col = min(numeric_columns)
        text_columns = [idx for idx in range(first_numeric_col)]
        if not text_columns:
            return None

        number_col = text_columns[0]
        organization_col = text_columns[-1]
        return {
            "column_count": column_count,
            "numeric_columns": numeric_columns,
            "first_numeric_col": first_numeric_col,
            "number_col": number_col,
            "organization_col": organization_col,
            "output_column_count": 2 + len(numeric_columns),
        }

    def _append_cell_text(self, row, index, value):
        if not value:
            return

        existing = row[index].strip()
        row[index] = f"{existing} {value}".strip() if existing else value

    def _compact_header_row(self, row, layout):
        compacted = [""] * layout["output_column_count"]
        text_values = []

        for col_idx in range(layout["first_numeric_col"]):
            if col_idx >= len(row):
                continue
            value = row[col_idx].strip()
            if value:
                text_values.append(value)

        for value in text_values:
            normalized = value.lower().strip()
            if normalized.startswith("no"):
                compacted[0] = value
                continue
            if "organisation" in normalized or "organization" in normalized:
                compacted[1] = value
                continue

            target_idx = 2 if layout["output_column_count"] > 2 else 1
            self._append_cell_text(compacted, target_idx, value)

        for target_idx, source_idx in enumerate(layout["numeric_columns"], start=2):
            if source_idx < len(row):
                self._append_cell_text(compacted, target_idx, row[source_idx].strip())

        return compacted

    def _compact_body_row(self, row, layout):
        compacted = [""] * layout["output_column_count"]
        detected_number = ""
        text_fragments = []

        for col_idx in range(layout["first_numeric_col"]):
            if col_idx >= len(row):
                continue

            value = row[col_idx].strip()
            if not value:
                continue

            if not detected_number and self._integer_like(value) and len(value.replace(",", "")) <= 3:
                detected_number = value
                continue

            match = re.match(r"^(\d{1,3})\s+(.+)$", value)
            if not detected_number and match:
                detected_number = match.group(1)
                value = match.group(2).strip()

            if value:
                text_fragments.append(value)

        if detected_number:
            compacted[0] = detected_number
        if text_fragments:
            compacted[1] = " ".join(text_fragments).strip()

        for target_idx, source_idx in enumerate(layout["numeric_columns"], start=2):
            if source_idx < len(row):
                compacted[target_idx] = row[source_idx].strip()

        return compacted

    def _realign_table_to_layout(self, header_rows, body_rows):
        layout = self._infer_body_layout(body_rows)
        if not layout:
            return header_rows, body_rows

        compacted_headers = [self._compact_header_row(row, layout) for row in header_rows]
        compacted_body = [self._compact_body_row(row, layout) for row in body_rows]
        return compacted_headers, compacted_body

    def _row_numeric_cells(self, row):
        return sum(1 for value in row if isinstance(value, str) and value.strip() and self._looks_numeric(value))

    def _row_text_cells(self, row):
        return [
            (idx, value.strip())
            for idx, value in enumerate(row)
            if isinstance(value, str) and value.strip() and not self._looks_numeric(value)
        ]

    def _row_missing_primary_text(self, row, layout):
        text_indexes = range(layout["first_numeric_col"]) if layout else range(len(row))
        for idx in text_indexes:
            if idx >= len(row):
                continue
            value = row[idx].strip()
            if value and not self._looks_numeric(value):
                return False
        return True

    def _merge_text_row_into_target(self, target_row, text_row, layout, prepend=False):
        merged = list(target_row)
        text_cells = self._row_text_cells(text_row)
        if not text_cells:
            return merged

        target_idx = layout["organization_col"] if layout else text_cells[0][0]
        if target_idx >= len(merged):
            merged.extend([""] * (target_idx + 1 - len(merged)))

        incoming = " ".join(value for _, value in text_cells).strip()
        existing = merged[target_idx].strip()
        if prepend:
            merged[target_idx] = f"{incoming} {existing}".strip() if existing else incoming
        else:
            merged[target_idx] = f"{existing} {incoming}".strip() if existing else incoming
        return merged

    def _merge_continuation_rows(self, rows, layout=None):
        if not rows:
            return rows

        merged_rows = []
        pending_rows = [list(row) for row in rows]
        idx = 0

        while idx < len(pending_rows):
            row = pending_rows[idx]
            is_text_only_sparse = self._is_section_like_row(row) and self._row_numeric_cells(row) == 0
            previous_row = merged_rows[-1] if merged_rows else None
            next_row = pending_rows[idx + 1] if idx + 1 < len(pending_rows) else None

            if is_text_only_sparse:
                if previous_row and self._row_numeric_cells(previous_row) >= 2 and self._row_missing_primary_text(previous_row, layout):
                    merged_rows[-1] = self._merge_text_row_into_target(previous_row, row, layout, prepend=False)
                    idx += 1
                    continue

                if next_row and self._row_numeric_cells(next_row) >= 2 and self._row_missing_primary_text(next_row, layout):
                    pending_rows[idx + 1] = self._merge_text_row_into_target(next_row, row, layout, prepend=True)
                    idx += 1
                    continue

            merged_rows.append(row)
            idx += 1

        return merged_rows

    def _normalize_selected_table(self, table):
        if not isinstance(table, dict):
            return table

        rows = table.get("rows", [])
        if not rows:
            return table

        header_row_count = min(table.get("header_row_count", 0), len(rows))
        repaired_rows, repaired_header_count, repaired_merges = _normalize_participating_org_table(rows, header_row_count)
        if repaired_rows != rows or repaired_header_count != header_row_count or repaired_merges:
            row_types = ["header"] * repaired_header_count + ["data"] * max(0, len(repaired_rows) - repaired_header_count)
            normalized = dict(table)
            normalized["rows"] = repaired_rows
            normalized["row_types"] = row_types
            normalized["header_row_count"] = repaired_header_count
            normalized["merges"] = repaired_merges
            normalized["row_count"] = len(repaired_rows)
            normalized["column_count"] = max((len(row) for row in repaired_rows), default=0)
            return normalized

        header_rows = rows[:header_row_count]
        body_rows = rows[header_row_count:]
        header_rows, body_rows = self._realign_table_to_layout(header_rows, body_rows)
        row_types = table.get("row_types", (["header"] * len(header_rows)) + self._classify_body_rows(body_rows))
        final_rows, row_types = self._dedupe_rows_with_types(header_rows + body_rows, row_types)
        header_row_count = 0
        for row_type in row_types:
            if row_type == "header":
                header_row_count += 1
            else:
                break

        normalized = dict(table)
        normalized["rows"] = final_rows
        normalized["row_types"] = row_types
        normalized["header_row_count"] = header_row_count
        normalized["merges"] = self._infer_header_merges_from_rows(final_rows, header_row_count)
        normalized["row_count"] = len(final_rows)
        normalized["column_count"] = max((len(row) for row in final_rows), default=0)
        return normalized

    def _normalize_selected_tables(self, tables):
        return [self._normalize_selected_table(table) for table in tables]

    def _numeric_count(self, row):
        return sum(1 for value in row if value.strip() and self._looks_numeric(value))

    def _extract_body_row_number(self, row, layout):
        for col_idx in range(layout["first_numeric_col"]):
            if col_idx >= len(row):
                continue

            value = row[col_idx].strip()
            if not value:
                continue

            if self._integer_like(value) and len(value.replace(",", "")) <= 3:
                return value

            match = re.match(r"^(\d{1,3})\s+.+$", value)
            if match:
                return match.group(1)

        return ""

    def _collapse_body_group(self, group_rows, layout):
        if not group_rows:
            return None

        collapsed = [""] * layout["column_count"]
        numeric_columns = layout["numeric_columns"]

        for col_idx in numeric_columns:
            for row in group_rows:
                if col_idx < len(row) and row[col_idx].strip():
                    collapsed[col_idx] = row[col_idx].strip()
                    break

        text_fragments = []
        detected_number = ""
        for row in group_rows:
            for col_idx in range(layout["first_numeric_col"]):
                if col_idx >= len(row):
                    continue
                value = row[col_idx].strip()
                if not value:
                    continue

                if not detected_number and self._integer_like(value) and len(value.replace(",", "")) <= 3:
                    detected_number = value
                    continue

                match = re.match(r"^(\d{1,3})\s+(.+)$", value)
                if not detected_number and match:
                    detected_number = match.group(1)
                    value = match.group(2).strip()

                if value:
                    text_fragments.append(value)

        if detected_number:
            collapsed[layout["number_col"]] = detected_number

        deduped_fragments = []
        for fragment in text_fragments:
            if not deduped_fragments or deduped_fragments[-1] != fragment:
                deduped_fragments.append(fragment)

        organization_value = " ".join(deduped_fragments).strip()
        if organization_value:
            collapsed[layout["organization_col"]] = organization_value

        return collapsed

    def _consolidate_body_rows(self, body_rows):
        layout = self._infer_body_layout(body_rows)
        if not layout:
            return body_rows

        body_rows = self._merge_continuation_rows(body_rows, layout)
        consolidated = []
        current_group = []
        current_group_has_numeric = False

        for row in body_rows:
            if self._is_section_like_row(row):
                if current_group:
                    collapsed = self._collapse_body_group(current_group, layout)
                    if collapsed:
                        consolidated.append(collapsed)
                consolidated.append(row)
                current_group = []
                current_group_has_numeric = False
                continue

            row_number = self._extract_body_row_number(row, layout)

            if row_number and current_group:
                collapsed = self._collapse_body_group(current_group, layout)
                if collapsed:
                    consolidated.append(collapsed)
                current_group = [row]
                current_group_has_numeric = self._numeric_count(row) >= 2
                continue

            row_has_numeric = self._numeric_count(row) >= 2
            if row_has_numeric and current_group and current_group_has_numeric:
                collapsed = self._collapse_body_group(current_group, layout)
                if collapsed:
                    consolidated.append(collapsed)
                current_group = [row]
                current_group_has_numeric = True
                continue

            current_group.append(row)
            current_group_has_numeric = current_group_has_numeric or row_has_numeric

        if current_group:
            collapsed = self._collapse_body_group(current_group, layout)
            if collapsed:
                consolidated.append(collapsed)

        return consolidated

    def _build_matrix(self, rows):
        column_centers = self._infer_column_centers(rows)
        if not column_centers:
            return [], []

        column_bounds = self._build_column_bounds(column_centers)
        matrix_rows = []
        merges = []

        for row_idx, row in enumerate(rows):
            values = [""] * len(column_centers)
            used_indexes = set()
            row_spans = []

            for item in row["items"]:
                column_idx = self._find_column_index(item, column_centers, column_bounds, used_indexes)
                values[column_idx] = item["text"]
                used_indexes.add(column_idx)

                covered_indexes = [
                    idx for idx, center in enumerate(column_centers)
                    if item["x"] - self.x_threshold <= center <= item["max_x"] + self.x_threshold
                ]
                if len(covered_indexes) >= 2:
                    row_spans.append((min(covered_indexes), max(covered_indexes)))

            matrix_rows.append(values)
            for start_idx, end_idx in row_spans:
                merges.append({
                    "row": row_idx,
                    "start_col": start_idx,
                    "end_col": end_idx,
                })

        non_empty_columns = [
            col_idx for col_idx in range(len(column_centers))
            if any(row[col_idx].strip() for row in matrix_rows)
        ]

        if not non_empty_columns:
            return [], []

        index_map = {old_idx: new_idx for new_idx, old_idx in enumerate(non_empty_columns)}
        normalized_rows = [
            [row[col_idx] for col_idx in non_empty_columns]
            for row in matrix_rows
        ]
        normalized_rows, _ = self._merge_sparse_columns(normalized_rows)
        normalized_rows = self._remove_duplicate_rows(normalized_rows)
        normalized_merges = []
        for merge in merges:
            covered = [
                index_map[idx]
                for idx in range(merge["start_col"], merge["end_col"] + 1)
                if idx in index_map
            ]
            if len(covered) >= 2:
                normalized_merges.append({
                    "row": merge["row"],
                    "start_col": min(covered),
                    "end_col": max(covered),
                })

        return self._normalize_organization_column(normalized_rows), normalized_merges

    def _looks_numeric(self, value):
        compact = value.replace(",", "").replace("%", "").replace("(", "").replace(")", "").strip()
        return bool(compact) and bool(re.fullmatch(r"[-+]?\d+(?:\.\d+)?", compact))

    def _infer_header_row_count(self, matrix_rows):
        if not matrix_rows:
            return 0

        for row_idx, row in enumerate(matrix_rows):
            non_empty_cells = [cell.strip() for cell in row if cell and cell.strip()]
            if not non_empty_cells:
                continue

            numeric_cells = sum(1 for cell in non_empty_cells if self._looks_numeric(cell))
            first_cell = non_empty_cells[0]
            starts_like_data = first_cell.isdigit() or numeric_cells >= max(2, len(non_empty_cells) // 2)

            if starts_like_data:
                return max(1, row_idx)

        return min(len(matrix_rows), 2)

    def _normalize_table(self, rows):
        trimmed_rows = self._trim_non_table_rows(self._remove_artifact_rows(rows))
        if not trimmed_rows:
            return None

        matrix_rows, merges = self._build_matrix(trimmed_rows)
        if not matrix_rows:
            return None

        return self._assemble_table(matrix_rows, self._consolidate_body_rows, fallback_merges=merges)

    def _detect_tables_fallback(self, items):
        rows = self._cluster_rows(items)
        table_groups = self._split_tables(rows)
        normalized_tables = []
        for table_rows in table_groups:
            table = self._normalize_table(table_rows)
            if table:
                normalized_tables.append(table)
        return normalized_tables
    
    def detect_tables(self, ocr_result):
        """
        Extract tables from OCR result.
        
        Args:
            ocr_result: List of [bbox, text, confidence] from PaddleOCR
            
        Returns:
            List of tables, each table is list of rows, each row is list of cell strings
        """
        if not ocr_result:
            return []
        
        items = self._extract_items(ocr_result)
        if not items:
            return []

        dynamic_tables = self._build_table_from_row_grid(self._remove_artifact_items(items))
        structured_tables = self._extract_structured_tables(items)
        fallback_tables = self._detect_tables_fallback(items)
        selected_tables = self._select_best_tables(dynamic_tables, structured_tables, fallback_tables)
        return self._normalize_selected_tables(selected_tables)


def normalize_ocr_result(ocr_result):
    """
    PaddleOCR returns a nested list per input image. For a single image we want
    the list of text boxes for that page.
    """
    if not ocr_result:
        return []

    if (
        isinstance(ocr_result, list)
        and ocr_result
        and isinstance(ocr_result[0], list)
        and ocr_result[0]
        and isinstance(ocr_result[0][0], list)
    ):
        return ocr_result[0]

    return ocr_result


def looks_numeric(value):
    if not isinstance(value, str):
        return isinstance(value, (int, float))

    compact = value.replace(",", "").replace("%", "").replace("(", "").replace(")", "").strip()
    return bool(compact) and bool(re.fullmatch(r"[-+]?\d+(?:\.\d+)?", compact))


def _normalized_text(value):
    return " ".join(str(value).strip().split()).lower() if value is not None else ""


def _table_header_similarity(previous_table, current_table):
    previous_rows = previous_table.get("rows", [])[: max(1, min(2, previous_table.get("header_row_count", 0)))]
    current_rows = current_table.get("rows", [])[: max(1, min(2, current_table.get("header_row_count", 0)))]
    previous_values = [_normalized_text(cell) for row in previous_rows for cell in row if _normalized_text(cell)]
    current_values = [_normalized_text(cell) for row in current_rows for cell in row if _normalized_text(cell)]

    if not previous_values or not current_values:
        return 0.0

    overlap = sum(1 for value in current_values if value in previous_values)
    return overlap / max(len(current_values), len(previous_values))


def _merge_continued_page_tables(pages):
    previous_table = None

    for page in pages:
        merged_tables = []
        for table in page.get("tables", []):
            if (
                previous_table
                and isinstance(previous_table, dict)
                and isinstance(table, dict)
                and previous_table.get("column_count") == table.get("column_count")
                and _table_header_similarity(previous_table, table) >= 0.8
            ):
                current_header_count = min(table.get("header_row_count", 0), len(table.get("rows", [])))
                previous_table["rows"].extend(table.get("rows", [])[current_header_count:])
                previous_types = previous_table.get("row_types", ["header"] * previous_table.get("header_row_count", 0))
                previous_table["row_types"] = previous_types + table.get("row_types", ["data"] * max(0, len(table.get("rows", [])) - current_header_count))[current_header_count:]
                previous_table["row_count"] = len(previous_table["rows"])
                previous_table["merges"] = []
                continue

            merged_tables.append(table)
            previous_table = table

        page["tables"] = merged_tables

    return pages


def _write_debug_payload(filename, payload):
    debug_dir = os.getenv("PDF_EXTRACT_DEBUG_DIR", "").strip()
    if not debug_dir:
        return

    target_dir = Path(debug_dir)
    target_dir.mkdir(parents=True, exist_ok=True)
    target_path = target_dir / filename
    target_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def _integer_like_text(value):
    compact = value.replace(",", "").strip()
    return bool(compact) and compact.isdigit()


def _append_export_value(row, index, value):
    if not value:
        return

    existing = row[index].strip()
    row[index] = f"{existing} {value}".strip() if existing else value


def _normalized_cell_text(value):
    return " ".join(str(value).strip().split()) if value is not None else ""


def _normalized_phrase(value):
    return _normalized_cell_text(value).lower()


def _dedupe_fragments(fragments):
    deduped = []
    for fragment in fragments:
        cleaned = _normalized_cell_text(fragment)
        if cleaned and (not deduped or deduped[-1] != cleaned):
            deduped.append(cleaned)
    return deduped


def _extract_row_number_from_text(value):
    cleaned = _normalized_cell_text(value)
    if not cleaned:
        return "", ""

    if _integer_like_text(cleaned) and len(cleaned.replace(",", "")) <= 3:
        return cleaned, ""

    match = re.match(r"^(\d{1,3})\s+(.+)$", cleaned)
    if match:
        return match.group(1), match.group(2).strip()

    return "", cleaned


def _is_known_header_row(row):
    normalized_cells = [_normalized_phrase(cell) for cell in row if _normalized_phrase(cell)]
    if not normalized_cells:
        return False

    joined = " ".join(normalized_cells)
    header_hits = 0
    header_tokens = [
        "participating organisation",
        "trading volume",
        "trading value",
        "year-to-date",
        "mar-26",
        "unit",
        "rm",
        "no.",
        "no",
    ]
    for token in header_tokens:
        if token in joined:
            header_hits += 1

    return header_hits >= 2


def _is_summary_row_name(value):
    normalized = _normalized_phrase(value)
    return normalized in {"total", "subtotal", "grand total"}


def _clean_participating_org_name(value):
    cleaned = _normalized_cell_text(value)
    cleaned = re.sub(r"^(participating organisation)\s+", "", cleaned, flags=re.IGNORECASE)
    return cleaned.strip()


def _looks_like_participating_org_layout(rows, header_row_count):
    if not rows:
        return False

    max_columns = max((len(row) for row in rows), default=0)
    if max_columns < 6:
        return False

    joined_head = " ".join(
        _normalized_phrase(cell)
        for row in rows[: max(3, min(len(rows), header_row_count + 2))]
        for cell in row
    )
    if "participating organisation" in joined_head:
        return True

    qualifying_rows = 0
    for row in rows[header_row_count:]:
        cells = [_normalized_cell_text(cell) for cell in row]
        text_like = sum(1 for value in cells[1:4] if value and not looks_numeric(value))
        numeric_like = sum(1 for value in cells if value and looks_numeric(value))
        if text_like >= 1 and numeric_like >= 4:
            qualifying_rows += 1

    return qualifying_rows >= 3


def _normalize_participating_org_table(rows, header_row_count):
    if not _looks_like_participating_org_layout(rows, header_row_count):
        return rows, header_row_count, []

    normalized_rows = []
    next_number = None

    for row in rows:
        if _is_known_header_row(row):
            continue

        cells = [_normalized_cell_text(cell) for cell in row]
        row_number = ""
        name_parts = []
        used_name_indexes = set()

        for col_idx in range(min(4, len(cells))):
            value = cells[col_idx]
            if not value:
                continue

            extracted_number, remainder = _extract_row_number_from_text(value)
            if col_idx == 0 and extracted_number and not row_number:
                row_number = extracted_number
                if remainder:
                    name_parts.append(remainder)
                continue

            if col_idx in (1, 2, 3) and not looks_numeric(value):
                used_name_indexes.add(col_idx)
                if extracted_number and not row_number:
                    row_number = extracted_number
                if remainder:
                    name_parts.append(remainder)
                continue

            if col_idx in (1, 2, 3) and extracted_number and not row_number:
                row_number = extracted_number

        name_parts = _dedupe_fragments(name_parts)
        metrics = []
        for col_idx, value in enumerate(cells):
            if not value or col_idx in used_name_indexes:
                continue
            if col_idx < 4 and row_number and value == row_number:
                continue
            if looks_numeric(value):
                metrics.append(value)

        metrics = metrics[-4:]
        while len(metrics) < 4:
            metrics.append("")

        name_value = _clean_participating_org_name(" ".join(name_parts).strip())
        if not name_value and not any(metrics):
            continue

        if not _is_summary_row_name(name_value):
            expected_number = 1 if next_number is None else next_number
            if not row_number:
                row_number = str(expected_number)
            elif _integer_like_text(row_number):
                parsed_number = int(row_number)
                if next_number is not None and parsed_number != expected_number:
                    row_number = str(expected_number)

        compacted_row = [row_number, name_value, metrics[0], metrics[1], metrics[2], metrics[3]]
        normalized_rows.append(compacted_row)

        if compacted_row[0] and _integer_like_text(compacted_row[0]):
            next_number = int(compacted_row[0]) + 1

    final_rows = [["No.", "Participating Organisation", "Unit", "%", "RM", "%"]] + normalized_rows
    return final_rows, 1, []


def _compact_table_for_export(rows, header_row_count):
    return rows, header_row_count, []


@app.get("/")
async def root():
    """Serve the main page"""
    return FileResponse(FRONTEND_DIR / "index.html", media_type="text/html")


@app.post("/upload")
async def upload_pdf(file: UploadFile = File(...)):
    """
    Upload and process a PDF file.
    Extract tables from all pages using PaddleOCR.
    
    Returns JSON with structure:
    {
        "status": "success",
        "pages": [
            {
                "page_number": 1,
                "tables": [
                    [["cell1", "cell2"], ["cell3", "cell4"]]
                ]
            }
        ]
    }
    """
    
    ocr_instance = get_ocr()
    if not ocr_instance:
        detail = OCR_INIT_ERROR or OCR_IMPORT_ERROR or "PaddleOCR not initialized"
        raise HTTPException(status_code=500, detail=detail)
    
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Only PDF files are allowed")
    
    temp_dir = None
    try:
        # Save uploaded file temporarily
        temp_dir = tempfile.mkdtemp()
        file_path = Path(temp_dir) / file.filename
        
        contents = await file.read()
        with open(file_path, 'wb') as f:
            f.write(contents)
        
        # Convert PDF pages to images
        images = convert_from_path(str(file_path), dpi=300)
        
        # Process each page
        pages_data = []
        detector = TableDetector(y_threshold=15)
        
        for page_num, image in enumerate(images, 1):
            image = image.filter(ImageFilter.SHARPEN)
            # Convert PIL Image to numpy array for PaddleOCR
            image_np = np.array(image)
            
            tables = detector.detect_tables_from_image(image_np, ocr_instance)
            _write_debug_payload(
                f"{Path(file.filename).stem}_page_{page_num}.json",
                {
                    "page_number": page_num,
                    "tables": tables,
                    "debug_snapshots": detector.debug_snapshots,
                },
            )
            detector.debug_snapshots = []
            
            pages_data.append({
                "page_number": page_num,
                "tables": tables
            })
        
        return {
            "status": "success",
            "pages": pages_data
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing PDF: {str(e)}")
    finally:
        if temp_dir and Path(temp_dir).exists():
            shutil.rmtree(temp_dir, ignore_errors=True)


@app.post("/download")
async def download_excel(data: dict):
    """
    Convert extracted table JSON to Excel and return as file download.
    
    Expected input:
    {
        "tables": [table1, table2, ...],
        "filename": "output.xlsx" (optional)
    }
    """
    try:
        tables = data.get("tables", [])
        filename = data.get("filename", "extracted_tables.xlsx")
        
        if not tables:
            raise HTTPException(status_code=400, detail="No tables provided")
        
        wb = Workbook()
        default_sheet = wb.active
        bold_font = Font(bold=True)

        for table_idx, table_entry in enumerate(tables, 1):
            if isinstance(table_entry, dict):
                table = table_entry.get("rows", [])
                page_number = table_entry.get("page_number")
                table_number = table_entry.get("table_number", table_idx)
                header_row_count = table_entry.get("header_row_count", 1)
                merges = table_entry.get("merges", [])
                row_types = table_entry.get("row_types", [])
            else:
                table = table_entry
                page_number = None
                table_number = table_idx
                header_row_count = 1
                merges = []
                row_types = []

            if not table:
                continue

            table, header_row_count, export_merges = _compact_table_for_export(table, header_row_count)
            if export_merges:
                merges = export_merges

            sheet_title = f"P{page_number or 1}_T{table_number}"
            ws = default_sheet if table_idx == 1 else wb.create_sheet()
            ws.title = sheet_title[:31]
            table_width = max((len(row) for row in table), default=1)

            for row_idx, row in enumerate(table):
                row_type = row_types[row_idx] if row_idx < len(row_types) else ("header" if row_idx < header_row_count else "data")
                if row_type == "section":
                    section_text = next((str(cell).strip() for cell in row if str(cell).strip()), "")
                    cell = ws.cell(row=row_idx + 1, column=1)
                    cell.value = section_text
                    cell.font = Font(bold=True, italic=True)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    if table_width > 1:
                        ws.merge_cells(
                            start_row=row_idx + 1,
                            start_column=1,
                            end_row=row_idx + 1,
                            end_column=table_width,
                        )
                    continue

                for col_idx, cell_value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx + 1, column=col_idx)
                    cell.value = cell_value
                    is_header = row_idx < header_row_count
                    is_numeric = looks_numeric(cell_value)

                    if is_header:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.font = bold_font
                    else:
                        cell.alignment = Alignment(
                            horizontal="right" if is_numeric else "left",
                            vertical="center"
                        )

            for merge in merges:
                merge_row = merge["row"] + 1
                start_col = merge["start_col"] + 1
                end_col = merge["end_col"] + 1
                if merge_row <= header_row_count and end_col > start_col:
                    ws.merge_cells(
                        start_row=merge_row,
                        start_column=start_col,
                        end_row=merge_row,
                        end_column=end_col,
                    )

            for column in ws.columns:
                max_length = 0
                first_real_cell = next((cell for cell in column if not isinstance(cell, MergedCell)), None)
                if first_real_cell is None:
                    continue
                column_letter = first_real_cell.column_letter
                for cell in column:
                    if isinstance(cell, MergedCell):
                        continue
                    try:
                        value_length = len(str(cell.value)) if cell.value is not None else 0
                        if value_length > max_length:
                            max_length = value_length
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        if len(wb.sheetnames) > 1 and default_sheet.title == "Sheet" and default_sheet.max_row == 1 and default_sheet["A1"].value is None:
            wb.remove(default_sheet)
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(temp_file.name)
        temp_file.close()
        
        return FileResponse(
            temp_file.name,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating Excel: {str(e)}")


@app.get("/health")
async def health_check():
    """Health check endpoint"""
    if ocr:
        ocr_status = "initialized"
    elif OCR_INIT_STARTED:
        ocr_status = "not_initialized"
    else:
        ocr_status = "not_initialized"

    response = {
        "status": "ok",
        "ocr_status": ocr_status
    }
    if OCR_IMPORT_ERROR:
        response["ocr_import_error"] = OCR_IMPORT_ERROR
    if OCR_INIT_ERROR:
        response["ocr_init_error"] = OCR_INIT_ERROR
    return response


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
